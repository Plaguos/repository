"""
table_append.py

⚠️ IMPORTANT
Ce fichier est une implémentation "raisonnable" pour l'append dans une Table Excel
(openpyxl) en auto-hébergement.

Si tu as déjà ton propre `table_append.py`, remplace celui-ci par le tien.

Hypothèses:
- Le classeur contient une table Excel (ListObject) nommée `table_name` dans `sheet_name`.
- On append des lignes à la fin de la table puis on étend le range de la table.
"""

from __future__ import annotations

from typing import List, Dict, Any
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table

def _find_table(ws, table_name: str) -> Table:
    for t in ws._tables.values():  # openpyxl stores tables here
        if getattr(t, "name", None) == table_name:
            return t
    # fallback: some versions keep ws.tables list
    for t in getattr(ws, "tables", {}).values() if isinstance(getattr(ws, "tables", None), dict) else getattr(ws, "tables", []):
        if getattr(t, "name", None) == table_name:
            return t
    raise ValueError(f"Table '{table_name}' introuvable dans la feuille '{ws.title}'.")

def _a1_to_coords(a1: str):
    # minimal A1 parser: e.g. "B3" -> (row=3, col=2)
    import re
    m = re.match(r"^([A-Z]+)(\d+)$", a1.upper())
    if not m:
        raise ValueError(f"Cellule A1 invalide: {a1}")
    col_letters, row_s = m.group(1), m.group(2)
    col = 0
    for ch in col_letters:
        col = col * 26 + (ord(ch) - ord("A") + 1)
    return int(row_s), col

def _coords_to_a1(row: int, col: int) -> str:
    # 1-indexed
    letters = ""
    n = col
    while n:
        n, r = divmod(n - 1, 26)
        letters = chr(r + ord("A")) + letters
    return f"{letters}{row}"

def append_rows_to_excel_table(
    workbook_path: str,
    sheet_name: str,
    table_name: str,
    rows: List[Dict[str, Any]],
) -> int:
    if not rows:
        return 0

    wb = load_workbook(workbook_path)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Feuille '{sheet_name}' introuvable dans le fichier.")
    ws = wb[sheet_name]

    table = _find_table(ws, table_name)

    # Table ref like "A1:M10"
    ref = table.ref
    start_a1, end_a1 = ref.split(":")
    start_row, start_col = _a1_to_coords(start_a1)
    end_row, end_col = _a1_to_coords(end_a1)

    # Header row is start_row, data starts at start_row+1
    header_row = start_row
    headers = [ws.cell(row=header_row, column=c).value for c in range(start_col, end_col + 1)]
    headers = [str(h).strip() if h is not None else "" for h in headers]

    # Append after the last row of the current table (end_row)
    write_row = end_row + 1
    for r in rows:
        for idx, h in enumerate(headers):
            col = start_col + idx
            if not h:
                ws.cell(row=write_row, column=col, value=None)
            else:
                ws.cell(row=write_row, column=col, value=r.get(h, ""))
        write_row += 1

    # Extend table ref down
    new_end_row = end_row + len(rows)
    table.ref = f"{_coords_to_a1(start_row, start_col)}:{_coords_to_a1(new_end_row, end_col)}"

    wb.save(workbook_path)
    return len(rows)
